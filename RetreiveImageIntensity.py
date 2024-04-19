
# -*- coding: utf-8 -*-
"""
Created on Wed Feb 14 11:28:11 2024

@author: SG
"""
import os
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook

filename=input("Please input the image´s name (without suffix):")
im=plt.imread(filename+".tif")
im1=im[:,:,1] # green channel

im45withlabel=im1[1150:1850,250:1050]
im_45withlabel=im1[100:800,1470:2170]

im45=im45withlabel[110:610,100:600]
im_45=im_45withlabel[130:630,100:600] # adjust these arraies to make the light in the center
plt.subplot(2, 2, 1)
plt.imshow(im45withlabel)
plt.title('+45')
plt.subplot(2, 2, 2)
plt.imshow(im_45withlabel)
plt.title('-45')

plt.subplot(2, 2, 3)
plt.imshow(im45)
plt.title('+45')
plt.subplot(2, 2, 4)
plt.imshow(im_45)
plt.title('-45')


def calIntensity(row, col,x,y,r,pol):
    count=0
    sumintensity=0 # for sum work here it should be 0
    for i in range(rows):  
        for j in range(cols):
            distance_squared = (i - x0) ** 2 + (j - y0) ** 2
            if distance_squared < r ** 2:
                if pol=="45":
                    sumintensity += im45[i, j]                    
                else:
                    sumintensity += im_45[i, j]
                count+=1
                return sumintensity/count
                    
    print(f"The sum of the intensity id {sumintensity} and the number of the pixel is {count} ")
    print(f"Then the average of the intensity is {sumintensity/count:.2f}")
    
print("For +45")
rows, cols = np.shape(im45)
x0,y0=rows/2,cols/2 
r=20
cal45=calIntensity(rows, cols, x0, y0, r,"45")
print("For -45")
rows_45, cols_45 = np.shape(im_45)
x0_45,y0_45=rows_45/2,cols_45/2 
cal_45=calIntensity(rows_45, cols_45, x0_45, y0_45, r,"-45")

excelName=input("please input a filename for saving data:")
if os.path.exists(excelName+ ".xlsx"):
    wb=load_workbook(excelName+r".xlsx")
    ws=wb.active
    ws.cell(int(float(filename)*10)+2,2,cal45)
    ws.cell(int(float(filename)*10)+2,3,cal_45)
    wb.save(excelName+r".xlsx")
else:
    wb=Workbook()
    ws=wb.active
    ws.cell(1,1,"Voltage")
    ws.cell(1,2,"pol=+45")
    ws.cell(1,3,"pol=-45")
    vols=np.arange(0,5.1,0.1)
    for k  in range(len(vols)):
          ws.cell(k+2, 1, vols[k])    
    ws.cell(int(float(filename)*10)+2,2,cal45)
    ws.cell(int(float(filename)*10)+2,3,cal_45)
    wb.save(excelName+r".xlsx")
    
















