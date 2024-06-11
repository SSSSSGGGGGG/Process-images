# -*- coding: utf-8 -*-
"""
Created on Wed Feb 14 11:28:11 2024

@author: SG
"""

import os
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook

os.chdir("C:/Users/Laboratorio/scitificCalculation/imageHub/b")
excelName=input("please input a filename for saving data:")
filename=0
maxinumber=16 # this value is for the upper bound in the reading imgae loop
while filename<maxinumber+1:
    
    im=plt.imread(str(filename*16)+".tif")
    im1=im[:,:,2] # green channel
    print(filename) 
    leftP=im1[0:2048,0:1223]
    rightP=im1[0:2048,1224:2448]

    im_0=leftP[565:625,540:600] #g=leftP[615:665,645:690], r=leftP[625:715,670:760]
    im_45=leftP[1590:1650,540:600] #g=leftP[1640:1690,645:690], r=leftP[1650:1740,670:760]

    im_n45=rightP[565:625,540:600] #b=rightP[565:625,540:600]
    im_90=rightP[1590:1650,540:600] #b=rightP[1590:1650,540:600]



    def calIntensity(row, col,x,y,r,pol):
        count=0
        sumintensity=0 # for sum work here it should be 0
        for i in range(row):  
            for j in range(col):
                distance_squared = (i - x0) ** 2 + (j - y0) ** 2
                if distance_squared < r ** 2:
                    if pol=="45":
                        sumintensity += im_45[i, j]                    
                    else:
                        sumintensity += im_n45[i, j]
                    count+=1
                    return sumintensity/count
                        
        print(f"The sum of the intensity id {sumintensity} and the number of the pixel is {count} ")
        print(f"Then the average of the intensity is {sumintensity/count:.2f}")
        
    # print("For +45")
    rows, cols = np.shape(im_45)
    x0,y0=rows/2,cols/2 
    r=20
    cal45=calIntensity(rows, cols, x0, y0, r,"45")
    # print("For -45")
    rows_45, cols_45 = np.shape(im_n45)
    x0_45,y0_45=rows_45/2,cols_45/2 
    cal_45=calIntensity(rows_45, cols_45, x0_45, y0_45, r,"-45")

    
    if os.path.exists(excelName+ ".xlsx"):
        wb=load_workbook(excelName+r".xlsx")
        ws=wb.active
        ws.cell(filename+2,2,cal45)
        ws.cell(filename+2,3,cal_45)
        wb.save(excelName+r".xlsx")
    else:
        wb=Workbook()
        ws=wb.active
        ws.cell(1,1,"ImageName")
        ws.cell(1,2,"pol=+45")
        ws.cell(1,3,"pol=-45")
        ws.cell(1,7,"Voltage")
        vols=np.arange(0,272,16)
        for k  in range(len(vols)):
              ws.cell(k+2, 1, vols[k])
              ws.cell(k+2, 7, vols[k])
        ws.cell(filename+2,2,cal45)
        ws.cell(filename+2,3,cal_45)
        wb.save(excelName+r".xlsx")
    filename+=1